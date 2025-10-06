#!/usr/bin/env python3
"""
Script to create a deployment-ready .xlsm file with VBA code embedded.

This script creates Ecom_Operations_Tracking_System.xlsm with all VBA code
pre-installed, so users can simply open and use it without manual import.

Requirements:
    - xlsxwriter (for creating Excel files with VBA support)
    - openpyxl (for reading existing Excel structure)
"""

import xlsxwriter
import openpyxl
import os
import sys
from datetime import datetime

def create_xlsm_with_vba():
    """
    Create a macro-enabled Excel workbook with VBA code embedded.
    
    This function:
    1. Reads the structure from the existing .xlsx file
    2. Creates a new .xlsm file with xlsxwriter
    3. Embeds the VBA project binary
    4. Copies all sheets and their data
    """
    
    print("=" * 60)
    print("E-Commerce Operations Tracking System")
    print("XLSM Generator with VBA")
    print("=" * 60)
    print()
    
    # File paths
    base_dir = os.path.dirname(os.path.abspath(__file__))
    source_xlsx = os.path.join(base_dir, 'Ecom_Operations_Tracking_System.xlsx')
    output_xlsm = os.path.join(base_dir, 'Ecom_Operations_Tracking_System.xlsm')
    vba_bin = os.path.join(base_dir, 'vbaProject.bin')
    vba_module = os.path.join(base_dir, 'VBA_Modules.bas')
    
    # Check if source files exist
    if not os.path.exists(source_xlsx):
        print(f"❌ Error: Source file not found: {source_xlsx}")
        return False
    
    if not os.path.exists(vba_module):
        print(f"❌ Error: VBA module not found: {vba_module}")
        return False
    
    print(f"📂 Reading structure from: {os.path.basename(source_xlsx)}")
    
    # Read the existing workbook structure
    try:
        wb_source = openpyxl.load_workbook(source_xlsx, data_only=False)
        print(f"✓ Loaded workbook with {len(wb_source.sheetnames)} sheets")
    except Exception as e:
        print(f"❌ Error loading workbook: {e}")
        return False
    
    # Create the new .xlsm workbook
    print(f"📝 Creating new macro-enabled workbook: {os.path.basename(output_xlsm)}")
    
    try:
        workbook = xlsxwriter.Workbook(output_xlsm, {'constant_memory': False})
        
        # Add VBA project if binary exists
        if os.path.exists(vba_bin):
            print(f"✓ Embedding VBA project from: {os.path.basename(vba_bin)}")
            workbook.add_vba_project(vba_bin)
        else:
            print(f"⚠ Warning: VBA binary not found: {os.path.basename(vba_bin)}")
            print(f"  The .xlsm will be created but VBA code must be added manually.")
            print(f"  See instructions in README_VBA.md")
        
        # Create formats for styling
        header_format = workbook.add_format({
            'bold': True,
            'bg_color': '#4472C4',
            'font_color': 'white',
            'border': 1
        })
        
        title_format = workbook.add_format({
            'bold': True,
            'font_size': 14,
            'bg_color': '#E7E6E6'
        })
        
        # Copy each sheet from source to destination
        for sheet_name in wb_source.sheetnames:
            print(f"  → Copying sheet: {sheet_name}")
            ws_source = wb_source[sheet_name]
            ws_dest = workbook.add_worksheet(sheet_name)
            
            # Copy cell values and basic formatting
            for row_idx, row in enumerate(ws_source.iter_rows()):
                for col_idx, cell in enumerate(row):
                    if cell.value is not None:
                        # Apply basic formatting for headers
                        if row_idx == 0 and sheet_name != 'Dashboard':
                            ws_dest.write(row_idx, col_idx, cell.value, header_format)
                        elif row_idx == 0 and sheet_name == 'Dashboard':
                            ws_dest.write(row_idx, col_idx, cell.value, title_format)
                        else:
                            ws_dest.write(row_idx, col_idx, cell.value)
            
            # Set column widths (approximate)
            if sheet_name == 'Dashboard':
                ws_dest.set_column('A:A', 40)
                ws_dest.set_column('B:E', 20)
            else:
                ws_dest.set_column('A:Z', 15)
        
        # Close the workbook
        workbook.close()
        print()
        print("✅ Successfully created macro-enabled workbook!")
        print(f"   File: {output_xlsm}")
        
        # File size
        if os.path.exists(output_xlsm):
            size_kb = os.path.getsize(output_xlsm) / 1024
            print(f"   Size: {size_kb:.1f} KB")
        
        return True
        
    except Exception as e:
        print(f"❌ Error creating workbook: {e}")
        import traceback
        traceback.print_exc()
        return False
    finally:
        wb_source.close()

def display_next_steps():
    """Display next steps for completing the VBA setup"""
    print()
    print("=" * 60)
    print("NEXT STEPS")
    print("=" * 60)
    print()
    print("If vbaProject.bin was not found, you need to create it:")
    print()
    print("Option 1 - Extract from existing .xlsm:")
    print("  1. If you have an .xlsm file with VBA code already")
    print("  2. Rename it to .zip")
    print("  3. Extract xl/vbaProject.bin")
    print("  4. Place vbaProject.bin in this directory")
    print("  5. Run this script again")
    print()
    print("Option 2 - Create manually in Excel:")
    print("  1. Open the generated .xlsm file")
    print("  2. Press Alt + F11 to open VBA Editor")
    print("  3. File → Import File → Select VBA_Modules.bas")
    print("  4. (Optional) Import BashQueryForm.frm")
    print("  5. Save and close")
    print()
    print("After VBA is embedded, users can:")
    print("  • Open the .xlsm file directly")
    print("  • Enable macros when prompted")
    print("  • Press Alt + F8 to run macros")
    print("  • Or click buttons if added to Dashboard")
    print()

def main():
    """Main entry point"""
    success = create_xlsm_with_vba()
    
    if success:
        display_next_steps()
        return 0
    else:
        print()
        print("❌ Failed to create .xlsm file")
        return 1

if __name__ == '__main__':
    sys.exit(main())
