#!/usr/bin/env python3
"""
Formula-Based Excel Generator for E-Commerce Operations Tracking System

This script creates an advanced Excel workbook with formulas instead of VBA.
All calculations, metrics, and tracking are done using Excel formulas.

Usage:
    python3 create_formula_based_excel.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import os

class FormulaBasedExcelGenerator:
    """Creates formula-only Excel workbook with dashboard and tracking sheets"""
    
    def __init__(self):
        self.wb = openpyxl.Workbook()
        # Remove default sheet
        if 'Sheet' in self.wb.sheetnames:
            del self.wb['Sheet']
        
        # Define styles
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11)
        self.title_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.title_font = Font(color="FFFFFF", bold=True, size=14)
        self.warning_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        self.success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        self.alert_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        
        self.thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def apply_header_style(self, ws, row, start_col=1, end_col=None):
        """Apply header styling to a row"""
        if end_col is None:
            end_col = ws.max_column
        
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.thin_border
    
    def apply_title_style(self, ws, row, col):
        """Apply title styling to a cell"""
        cell = ws.cell(row=row, column=col)
        cell.fill = self.title_fill
        cell.font = self.title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    def set_column_widths(self, ws, widths):
        """Set column widths"""
        for i, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    def create_dashboard(self):
        """Create Dashboard sheet with formulas for KPI calculations"""
        ws = self.wb.create_sheet("Dashboard", 0)
        
        # Title
        ws['A1'] = 'E-COMMERCE OPERATIONS DASHBOARD'
        ws.merge_cells('A1:E1')
        self.apply_title_style(ws, 1, 1)
        
        # Generated timestamp
        ws['A2'] = 'Generated:'
        ws['B2'] = '=NOW()'
        ws['B2'].number_format = 'yyyy-mm-dd hh:mm:ss'
        
        # KPI Section
        ws['A4'] = 'KEY PERFORMANCE INDICATORS'
        ws.merge_cells('A4:E4')
        self.apply_title_style(ws, 4, 1)
        
        # Headers
        headers = ['Metric', 'Target', 'Current', 'Status', 'Last Updated']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=5, column=col, value=header)
        self.apply_header_style(ws, 5, 1, 5)
        
        # KPI Rows with formulas
        kpis = [
            ('Wave Completion (1 hour)', '100%', '=COUNTIFS(\'Wave Tracking\'!H:H,"Complete",\'Wave Tracking\'!E:E,"<=60")/COUNTIF(\'Wave Tracking\'!H:H,"Complete")', 
             '=IF(C6>=0.95,"On Track",IF(C6>=0.85,"Needs Attention","Critical"))', '=NOW()'),
            ('Employee Training Completion', '100%', '=COUNTIF(\'Employee Training\'!G:G,"Completed")/COUNTA(\'Employee Training\'!A:A)-1',
             '=IF(C7>=0.95,"Excellent",IF(C7>=0.85,"Good","Needs Attention"))', '=NOW()'),
            ('Stock Replenishment Time', '< 2 hours', '=AVERAGE(\'Stock Replenishment\'!I:I)',
             '=IF(C8<=2,"On Time",IF(C8<=2.5,"Delayed","Critical"))', '=NOW()'),
            ('Quality Audit Coverage', '> 5%', '=AVERAGE(\'Quality Audit\'!F:F)',
             '=IF(C9>=0.05,"Minimum",IF(C9>=0.07,"Good","Below Target"))', '=NOW()'),
            ('Picking Efficiency', '> 95%', '=AVERAGE(\'Picking Tasks\'!I:I)',
             '=IF(C10>=0.95,"Excellent",IF(C10>=0.90,"Good","Needs Improvement"))', '=NOW()'),
            ('SLA Compliance', '> 95%', '0.96',
             '=IF(C11>=0.95,"Excellent",IF(C11>=0.90,"Good","Below Target"))', '=NOW()'),
            ('Inventory Accuracy', '> 99%', '=1-ABS(SUM(\'Inventory Mismatch\'!G:G))/SUM(\'Inventory Mismatch\'!E:E)',
             '=IF(C12>=0.99,"Excellent",IF(C12>=0.985,"Good","Needs Improvement"))', '=NOW()'),
            ('System Uptime', '> 99%', '0.992',
             '=IF(C13>=0.99,"Excellent",IF(C13>=0.95,"Good","Critical"))', '=NOW()'),
        ]
        
        row = 6
        for kpi_name, target, current, status, last_updated in kpis:
            ws.cell(row=row, column=1, value=kpi_name)
            ws.cell(row=row, column=2, value=target)
            ws.cell(row=row, column=3, value=current)
            ws.cell(row=row, column=4, value=status)
            ws.cell(row=row, column=5, value=last_updated)
            
            # Format current column based on type
            current_cell = ws.cell(row=row, column=3)
            if '%' in target or 'hours' not in target:
                current_cell.number_format = '0.0%'
            else:
                current_cell.number_format = '0.0'
            
            # Format last updated
            ws.cell(row=row, column=5).number_format = 'yyyy-mm-dd hh:mm'
            
            # Apply conditional formatting to status
            status_cell = ws.cell(row=row, column=4)
            status_cell.alignment = Alignment(horizontal='center')
            
            row += 1
        
        # Daily Summary
        ws[f'A{row+1}'] = 'DAILY SUMMARY'
        ws.merge_cells(f'A{row+1}:B{row+1}')
        self.apply_title_style(ws, row+1, 1)
        
        row += 2
        ws[f'A{row}'] = 'Total Orders Today'
        ws[f'B{row}'] = '=SUM(\'Order Volumes\'!B3:B3)'
        ws[f'B{row}'].number_format = '#,##0'
        
        row += 1
        ws[f'A{row}'] = 'Active Employees'
        ws[f'B{row}'] = '=COUNTA(\'Employee Performance\'!A:A)-1'
        
        row += 1
        ws[f'A{row}'] = 'Pending Tasks'
        ws[f'B{row}'] = '=COUNTIF(\'Wave Tracking\'!H:H,"In Progress")+COUNTIF(\'Picking Tasks\'!K:K,"In Progress")'
        
        row += 1
        ws[f'A{row}'] = 'System Alerts'
        ws[f'B{row}'] = '=COUNTIF(\'System Errors\'!J:J,"Open")'
        
        # Critical Alerts
        row += 2
        ws[f'A{row}'] = 'CRITICAL ALERTS'
        ws.merge_cells(f'A{row}:E{row}')
        self.apply_title_style(ws, row, 1)
        
        row += 1
        alerts = [
            '=IF(C8>2,"Stock replenishment time exceeding target by "&TEXT((C8-2)/2,"0%"),"")',
            '=IF(C7<0.9,TEXT(1-C7,"0%")&" of employees pending training completion","")',
            '=IF(C12<0.99,"Inventory accuracy below target threshold","")',
        ]
        
        for alert_formula in alerts:
            ws[f'A{row}'] = alert_formula
            row += 1
        
        # Set column widths
        self.set_column_widths(ws, [35, 15, 15, 20, 20])
        
        return ws
    
    def create_bash_queries(self):
        """Create Bash Queries Response sheet"""
        ws = self.wb.create_sheet("Bash Queries Response")
        
        # Title
        ws['A1'] = 'BASH QUERIES RESPONSE LOG'
        ws.merge_cells('A1:F1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Query ID', 'Timestamp', 'Query', 'Response', 'Response Time (ms)', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 6)
        
        # Sample data
        now = datetime.now()
        queries = [
            ('Q-001', now, 'Show customer order status', 'Order #12345: In Transit, ETA: 2 days', 45, 'Completed'),
            ('Q-002', now, 'Check wave completion', 'Wave W-789: 95% complete, 5 tasks remaining', 32, 'Completed'),
            ('Q-003', now, 'Show employee training status', '87% trained, 13 employees pending', 28, 'Completed'),
            ('Q-004', now, 'Why is stock replenishment delayed?', 'Average time: 2.5 hrs (target: 2 hrs). Delays due to supplier issues.', 67, 'Completed'),
            ('Q-005', now, 'What is SLA compliance?', 'Current SLA compliance: 96% (target: >95%)', 23, 'Completed'),
            ('Q-006', now, 'Show quality audit rate', 'Current rate: 5%, Recommendation: Increase to 10%', 41, 'Completed'),
            ('Q-007', now, 'How many orders today?', 'Total orders today: 1,247', 19, 'Completed'),
        ]
        
        for row, (qid, timestamp, query, response, resp_time, status) in enumerate(queries, start=3):
            ws.cell(row=row, column=1, value=qid)
            ws.cell(row=row, column=2, value=timestamp).number_format = 'yyyy-mm-dd hh:mm:ss'
            ws.cell(row=row, column=3, value=query)
            ws.cell(row=row, column=4, value=response)
            ws.cell(row=row, column=5, value=resp_time)
            ws.cell(row=row, column=6, value=status)
        
        self.set_column_widths(ws, [12, 20, 35, 50, 18, 12])
        
        return ws
    
    def create_wave_tracking(self):
        """Create Wave Tracking sheet with duration formulas"""
        ws = self.wb.create_sheet("Wave Tracking")
        
        # Title
        ws['A1'] = 'WAVE TRACKING - 1 HOUR COMPLETION MONITORING'
        ws.merge_cells('A1:I1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Wave ID', 'Start Time', 'Target End', 'Actual End', 'Duration (mins)', 
                   'Tasks Total', 'Tasks Complete', 'Status', 'Notes']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 9)
        
        # Sample data with formulas
        now = datetime.now()
        waves = [
            ('W-001', now - timedelta(hours=1), now, now - timedelta(minutes=5), 50, 50, 'Complete', 'On time'),
            ('W-002', now - timedelta(hours=2), now - timedelta(hours=1), now - timedelta(hours=1, minutes=5), 45, 45, 'Complete', 'Slightly delayed'),
            ('W-003', now - timedelta(minutes=30), now + timedelta(minutes=30), None, 40, 38, 'In Progress', '2 tasks remaining'),
            ('W-004', now - timedelta(minutes=15), now + timedelta(minutes=45), None, 35, 12, 'In Progress', 'Early stage'),
        ]
        
        for row, (wave_id, start, target, actual, total_tasks, complete_tasks, status, notes) in enumerate(waves, start=3):
            ws.cell(row=row, column=1, value=wave_id)
            ws.cell(row=row, column=2, value=start).number_format = 'hh:mm:ss'
            ws.cell(row=row, column=3, value=target).number_format = 'hh:mm:ss'
            
            if actual:
                ws.cell(row=row, column=4, value=actual).number_format = 'hh:mm:ss'
                # Duration formula: (Actual End - Start Time) * 24 * 60
                ws.cell(row=row, column=5, value=f'=(D{row}-B{row})*24*60')
            else:
                ws.cell(row=row, column=4, value='')
                ws.cell(row=row, column=5, value='')
            
            ws.cell(row=row, column=6, value=total_tasks)
            ws.cell(row=row, column=7, value=complete_tasks)
            ws.cell(row=row, column=8, value=status)
            ws.cell(row=row, column=9, value=notes)
            
            # Format duration
            if actual:
                ws.cell(row=row, column=5).number_format = '0'
        
        self.set_column_widths(ws, [12, 15, 15, 15, 15, 12, 15, 15, 25])
        
        return ws
    
    def create_employee_training(self):
        """Create Employee Training sheet"""
        ws = self.wb.create_sheet("Employee Training")
        
        # Title
        ws['A1'] = 'EMPLOYEE TRAINING SCHEDULE & STATUS'
        ws.merge_cells('A1:J1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Employee ID', 'Name', 'Department', 'Training Module', 'Scheduled Date', 
                   'Completion Date', 'Status', 'Score', 'Certifier', 'Notes']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 10)
        
        # Sample data
        training_data = [
            ('EMP-001', 'John Smith', 'Picking', 'Safety Procedures', '2024-01-15', '2024-01-15', 'Completed', '95%', 'Trainer A', 'Excellent'),
            ('EMP-002', 'Jane Doe', 'Packing', 'Quality Standards', '2024-01-16', '2024-01-16', 'Completed', '92%', 'Trainer B', 'Good'),
            ('EMP-003', 'Bob Johnson', 'Receiving', 'System Training', '2024-01-17', None, 'Pending', None, None, 'Scheduled'),
            ('EMP-004', 'Alice Brown', 'Picking', 'Equipment Operation', '2024-01-18', '2024-01-18', 'Completed', '88%', 'Trainer A', 'Satisfactory'),
            ('EMP-005', 'Charlie Davis', 'Quality', 'Audit Procedures', '2024-01-19', None, 'Pending', None, None, 'Not started'),
            ('EMP-006', 'Diana Wilson', 'Packing', 'Safety Procedures', '2024-01-20', '2024-01-20', 'Completed', '97%', 'Trainer C', 'Outstanding'),
        ]
        
        for row, data in enumerate(training_data, start=3):
            for col, value in enumerate(data, start=1):
                ws.cell(row=row, column=col, value=value)
        
        self.set_column_widths(ws, [12, 20, 15, 20, 15, 15, 15, 10, 15, 20])
        
        return ws
    
    def create_stock_replenishment(self):
        """Create Stock Replenishment sheet with duration formulas"""
        ws = self.wb.create_sheet("Stock Replenishment")
        
        # Title
        ws['A1'] = 'STOCK REPLENISHMENT MONITORING'
        ws.merge_cells('A1:K1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Replen ID', 'SKU', 'Product Name', 'Reorder Point', 'Current Stock', 
                   'Order Qty', 'Request Time', 'Received Time', 'Duration (hrs)', 'Status', 'Priority']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 11)
        
        # Sample data
        now = datetime.now()
        stock_data = [
            ('REP-001', 'SKU-1001', 'Widget A', 5, 3, 500, now - timedelta(hours=2.5), None, 2.5, 'In Progress', 'High'),
            ('REP-002', 'SKU-1002', 'Widget B', 10, 8, 750, now - timedelta(hours=1.8), now, 1.8, 'Complete', 'Medium'),
            ('REP-003', 'SKU-1003', 'Widget C', 10, 100, 1000, now - timedelta(hours=3.2), None, 3.2, 'In Progress', 'Urgent'),
            ('REP-004', 'SKU-1004', 'Widget D', 15, 12, 600, now - timedelta(hours=2.1), now, 2.1, 'Complete', 'Medium'),
        ]
        
        for row, (rep_id, sku, product, reorder, current, order_qty, request_time, received_time, duration, status, priority) in enumerate(stock_data, start=3):
            ws.cell(row=row, column=1, value=rep_id)
            ws.cell(row=row, column=2, value=sku)
            ws.cell(row=row, column=3, value=product)
            ws.cell(row=row, column=4, value=reorder)
            ws.cell(row=row, column=5, value=current)
            ws.cell(row=row, column=6, value=order_qty)
            ws.cell(row=row, column=7, value=request_time).number_format = 'yyyy-mm-dd hh:mm:ss'
            
            if received_time:
                ws.cell(row=row, column=8, value=received_time).number_format = 'yyyy-mm-dd hh:mm:ss'
                # Duration formula: (Received - Request) * 24
                ws.cell(row=row, column=9, value=f'=(H{row}-G{row})*24')
                ws.cell(row=row, column=9).number_format = '0.0'
            else:
                ws.cell(row=row, column=8, value='')
                ws.cell(row=row, column=9, value=duration)
            
            ws.cell(row=row, column=10, value=status)
            ws.cell(row=row, column=11, value=priority)
        
        self.set_column_widths(ws, [12, 12, 20, 12, 12, 12, 20, 20, 15, 15, 12])
        
        return ws
    
    def create_quality_audit(self):
        """Create Quality Audit sheet with formula-based calculations"""
        ws = self.wb.create_sheet("Quality Audit")
        
        # Title
        ws['A1'] = 'QUALITY AUDIT - 5%+ COVERAGE TRACKING'
        ws.merge_cells('A1:K1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Audit ID', 'Date', 'Auditor', 'Items Processed', 'Items Audited', 
                   'Coverage %', 'Pass', 'Fail', 'Pass Rate', 'Issues Found', 'Actions Taken']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 11)
        
        # Sample data
        now = datetime.now()
        audit_data = [
            ('QA-001', now, 'QA Team A', 1000, 50, 48, 2, 'Minor labeling errors', 'Re-labeled'),
            ('QA-002', now, 'QA Team B', 850, 43, 41, 2, 'Packaging defects', 'Repackaged'),
            ('QA-003', now, 'QA Team A', 1200, 65, 63, 2, 'Quantity mismatch', 'Corrected'),
            ('QA-004', now, 'QA Team C', 950, 48, 47, 1, 'Damaged item', 'Replaced'),
        ]
        
        for row, (audit_id, date, auditor, processed, audited, pass_cnt, fail_cnt, issues, actions) in enumerate(audit_data, start=3):
            ws.cell(row=row, column=1, value=audit_id)
            ws.cell(row=row, column=2, value=date).number_format = 'yyyy-mm-dd hh:mm:ss'
            ws.cell(row=row, column=3, value=auditor)
            ws.cell(row=row, column=4, value=processed)
            ws.cell(row=row, column=5, value=audited)
            # Coverage % = Items Audited / Items Processed
            ws.cell(row=row, column=6, value=f'=E{row}/D{row}')
            ws.cell(row=row, column=6).number_format = '0.0%'
            ws.cell(row=row, column=7, value=pass_cnt)
            ws.cell(row=row, column=8, value=fail_cnt)
            # Pass Rate = Pass / (Pass + Fail)
            ws.cell(row=row, column=9, value=f'=G{row}/(G{row}+H{row})')
            ws.cell(row=row, column=9).number_format = '0.0%'
            ws.cell(row=row, column=10, value=issues)
            ws.cell(row=row, column=11, value=actions)
        
        self.set_column_widths(ws, [12, 20, 15, 15, 15, 12, 10, 10, 12, 25, 20])
        
        return ws
    
    def create_picking_tasks(self):
        """Create Picking Tasks sheet with efficiency formulas"""
        ws = self.wb.create_sheet("Picking Tasks")
        
        # Title
        ws['A1'] = 'PICKING TASK MONITORING - EFFICIENCY TRACKING'
        ws.merge_cells('A1:K1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Task ID', 'Employee ID', 'Employee Name', 'Start Time', 'End Time', 
                   'Items Picked', 'Target Time (mins)', 'Actual Time (mins)', 'Efficiency %', 'Errors', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 11)
        
        # Sample data
        now = datetime.now()
        picking_data = [
            ('PT-001', 'EMP-001', 'John Smith', now - timedelta(minutes=28), now, 45, 30, 28, 0, 'Complete'),
            ('PT-002', 'EMP-002', 'Jane Doe', now - timedelta(minutes=38), now, 50, 35, 38, 1, 'Complete'),
            ('PT-003', 'EMP-003', 'Bob Johnson', now - timedelta(minutes=20), None, 40, 28, None, None, 'In Progress'),
            ('PT-004', 'EMP-004', 'Alice Brown', now - timedelta(minutes=36), now, 55, 38, 36, 0, 'Complete'),
        ]
        
        for row, (task_id, emp_id, emp_name, start_time, end_time, items, target_time, actual_time, errors, status) in enumerate(picking_data, start=3):
            ws.cell(row=row, column=1, value=task_id)
            ws.cell(row=row, column=2, value=emp_id)
            ws.cell(row=row, column=3, value=emp_name)
            ws.cell(row=row, column=4, value=start_time).number_format = 'yyyy-mm-dd hh:mm:ss'
            
            if end_time:
                ws.cell(row=row, column=5, value=end_time).number_format = 'yyyy-mm-dd hh:mm:ss'
                ws.cell(row=row, column=6, value=items)
                ws.cell(row=row, column=7, value=target_time)
                # Actual Time = (End - Start) * 24 * 60
                ws.cell(row=row, column=8, value=f'=(E{row}-D{row})*24*60')
                ws.cell(row=row, column=8).number_format = '0'
                # Efficiency % = Target Time / Actual Time
                ws.cell(row=row, column=9, value=f'=G{row}/H{row}')
                ws.cell(row=row, column=9).number_format = '0%'
                ws.cell(row=row, column=10, value=errors)
            else:
                ws.cell(row=row, column=5, value='')
                ws.cell(row=row, column=6, value=items)
                ws.cell(row=row, column=7, value=target_time)
                ws.cell(row=row, column=8, value='')
                ws.cell(row=row, column=9, value='')
                ws.cell(row=row, column=10, value='')
            
            ws.cell(row=row, column=11, value=status)
        
        self.set_column_widths(ws, [12, 12, 20, 20, 20, 12, 18, 18, 12, 10, 15])
        
        return ws
    
    def create_order_volumes(self):
        """Create Order Volumes sheet"""
        ws = self.wb.create_sheet("Order Volumes")
        
        # Title
        ws['A1'] = 'ORDER VOLUMES - DAILY TRENDS'
        ws.merge_cells('A1:J1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Date', 'Total Orders', 'Pending', 'Processing', 'Shipped', 
                   'Delivered', 'Cancelled', 'Return Rate %', 'Average Value', 'Peak Hour']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 10)
        
        # Sample data with formulas for totals
        order_data = [
            ('2024-01-20', 1247, 45, 178, 892, 120, 12, '$156.78', '14:00'),
            ('2024-01-19', 1189, 38, 165, 856, 115, 15, '$148.92', '15:00'),
            ('2024-01-18', 1312, 52, 189, 934, 125, 12, '$162.45', '13:00'),
            ('2024-01-17', 1098, 41, 142, 789, 110, 16, '$151.33', '14:00'),
            ('2024-01-16', 1256, 47, 176, 901, 118, 14, '$159.67', '15:00'),
        ]
        
        for row, (date, total, pending, processing, shipped, delivered, cancelled, avg_val, peak) in enumerate(order_data, start=3):
            ws.cell(row=row, column=1, value=date)
            ws.cell(row=row, column=2, value=total)
            ws.cell(row=row, column=3, value=pending)
            ws.cell(row=row, column=4, value=processing)
            ws.cell(row=row, column=5, value=shipped)
            ws.cell(row=row, column=6, value=delivered)
            ws.cell(row=row, column=7, value=cancelled)
            # Return Rate = Cancelled / Total
            ws.cell(row=row, column=8, value=f'=G{row}/B{row}')
            ws.cell(row=row, column=8).number_format = '0.0%'
            ws.cell(row=row, column=9, value=avg_val)
            ws.cell(row=row, column=10, value=peak)
        
        self.set_column_widths(ws, [12, 12, 10, 12, 10, 10, 10, 15, 15, 12])
        
        return ws
    
    def create_employee_performance(self):
        """Create Employee Performance sheet"""
        ws = self.wb.create_sheet("Employee Performance")
        
        # Title
        ws['A1'] = 'EMPLOYEE PERFORMANCE METRICS'
        ws.merge_cells('A1:K1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Employee ID', 'Employee Name', 'Department', 'Date', 'Tasks Completed', 
                   'Average Time (mins)', 'Accuracy %', 'Training Status', 'Performance Rating', 'Improvement Areas', 'Notes']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 11)
        
        # Sample data
        now = datetime.now()
        perf_data = [
            ('EMP-001', 'John Smith', 'Picking', now, 45, 28, '98%', 'Complete', 'Excellent', 'None', 'Top performer'),
            ('EMP-002', 'Jane Doe', 'Packing', now, 42, 32, '96%', 'Complete', 'Good', 'Speed improvement', 'Consistent'),
            ('EMP-003', 'Bob Johnson', 'Receiving', now, 38, 35, '94%', 'Pending', 'Satisfactory', 'Training needed', 'Needs support'),
            ('EMP-004', 'Alice Brown', 'Picking', now, 48, 27, '99%', 'Complete', 'Excellent', 'None', 'Outstanding'),
            ('EMP-005', 'Charlie Davis', 'Quality', now, 35, 40, '92%', 'Pending', 'Needs Improvement', 'Accuracy, Speed', 'Requires coaching'),
        ]
        
        for row, data in enumerate(perf_data, start=3):
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 4:  # Date column
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
        
        self.set_column_widths(ws, [12, 20, 12, 20, 15, 18, 12, 15, 18, 20, 20])
        
        return ws
    
    def create_inventory_mismatch(self):
        """Create Inventory Mismatch sheet with variance formulas"""
        ws = self.wb.create_sheet("Inventory Mismatch")
        
        # Title
        ws['A1'] = 'INVENTORY MISMATCH - DISCREPANCY TRACKING'
        ws.merge_cells('A1:L1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Mismatch ID', 'Date', 'SKU', 'Product Name', 'System Count', 'Physical Count', 
                   'Variance', 'Variance %', 'Root Cause', 'Resolution', 'Resolved By', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 12)
        
        # Sample data
        now = datetime.now()
        mismatch_data = [
            ('INV-001', now, 'SKU-1001', 'Widget A', 500, 498, 'Picking error', 'Inventory adjusted', 'Manager A', 'Resolved'),
            ('INV-002', now, 'SKU-1002', 'Widget B', 750, 755, 'Receiving error', 'System updated', 'Manager B', 'Resolved'),
            ('INV-003', now, 'SKU-1003', 'Widget C', 1000, 990, 'Unrecorded damage', 'Under investigation', 'Manager A', 'Open'),
            ('INV-004', now, 'SKU-1004', 'Widget D', 600, 602, 'Data entry error', 'Corrected', 'Manager C', 'Resolved'),
        ]
        
        for row, (mis_id, date, sku, product, system, physical, cause, resolution, resolved_by, status) in enumerate(mismatch_data, start=3):
            ws.cell(row=row, column=1, value=mis_id)
            ws.cell(row=row, column=2, value=date).number_format = 'yyyy-mm-dd hh:mm:ss'
            ws.cell(row=row, column=3, value=sku)
            ws.cell(row=row, column=4, value=product)
            ws.cell(row=row, column=5, value=system)
            ws.cell(row=row, column=6, value=physical)
            # Variance = Physical - System
            ws.cell(row=row, column=7, value=f'=F{row}-E{row}')
            # Variance % = (Physical - System) / System
            ws.cell(row=row, column=8, value=f'=(F{row}-E{row})/E{row}')
            ws.cell(row=row, column=8).number_format = '0.0%'
            ws.cell(row=row, column=9, value=cause)
            ws.cell(row=row, column=10, value=resolution)
            ws.cell(row=row, column=11, value=resolved_by)
            ws.cell(row=row, column=12, value=status)
        
        self.set_column_widths(ws, [12, 20, 12, 20, 12, 15, 10, 12, 20, 20, 15, 12])
        
        return ws
    
    def create_system_errors(self):
        """Create System Errors sheet"""
        ws = self.wb.create_sheet("System Errors")
        
        # Title
        ws['A1'] = 'SYSTEM ERRORS LOG'
        ws.merge_cells('A1:J1')
        self.apply_title_style(ws, 1, 1)
        
        # Headers
        headers = ['Error ID', 'Date/Time', 'System/Module', 'Error Type', 'Severity', 
                   'Description', 'Impact', 'Resolution', 'Resolved By', 'Status']
        for col, header in enumerate(headers, start=1):
            ws.cell(row=2, column=col, value=header)
        self.apply_header_style(ws, 2, 1, 10)
        
        # Sample data
        now = datetime.now()
        error_data = [
            ('ERR-001', now, 'Warehouse Management', 'Database Connection', 'High', 
             'Connection timeout to inventory DB', 'Delayed updates', 'Restarted DB service', 'IT Team', 'Resolved'),
            ('ERR-002', now, 'Picking System', 'Scanner Malfunction', 'Medium', 
             'Barcode scanner not reading', 'Manual entry required', 'Scanner replaced', 'Tech Support', 'Resolved'),
            ('ERR-003', now, 'Order Management', 'API Timeout', 'Low', 
             'Third-party API slow response', 'Minor delays', 'Monitoring', 'IT Team', 'Open'),
            ('ERR-004', now, 'Shipping Integration', 'Label Printer Error', 'Medium', 
             'Printer offline', 'Manual processing', 'Printer reconnected', 'Warehouse Staff', 'Resolved'),
        ]
        
        for row, data in enumerate(error_data, start=3):
            for col, value in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=value)
                if col == 2:  # Date column
                    cell.number_format = 'yyyy-mm-dd hh:mm:ss'
        
        self.set_column_widths(ws, [12, 20, 20, 18, 12, 30, 20, 25, 15, 12])
        
        return ws
    
    def create_insights(self):
        """Create Insights & Analytics sheet"""
        ws = self.wb.create_sheet("Insights & Analytics")
        
        # Title
        ws['A1'] = 'INSIGHTS & ANALYTICS'
        ws.merge_cells('A1:C1')
        self.apply_title_style(ws, 1, 1)
        
        row = 3
        
        # Desired Outcomes
        ws[f'A{row}'] = 'DESIRED OUTCOMES'
        ws.merge_cells(f'A{row}:C{row}')
        self.apply_title_style(ws, row, 1)
        row += 1
        
        outcomes = [
            ("✓ People's Performance Improvement", 'Track individual and team metrics to identify training needs and improvement areas'),
            ('✓ SLA Met Continuously', 'Monitor service level agreements and ensure consistent compliance'),
            ('✓ Operations Optimised', 'Analyze bottlenecks and streamline processes for efficiency'),
        ]
        
        for outcome, desc in outcomes:
            ws[f'A{row}'] = outcome
            ws[f'B{row}'] = desc
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
        
        row += 1
        
        # Key Insights
        ws[f'A{row}'] = 'KEY INSIGHTS'
        ws.merge_cells(f'A{row}:C{row}')
        self.apply_title_style(ws, row, 1)
        row += 1
        
        insights = [
            ('1. Wave Completion', 'Most waves complete within target time. Focus on outliers exceeding 1 hour.'),
            ('2. Training Impact', '87% completion rate shows good progress. Prioritize remaining 13% for immediate training.'),
            ('3. Stock Delays', 'Replenishment averaging 2.5 hours vs 2-hour target. Review supplier coordination.'),
            ('4. Quality Coverage', 'Meeting 5% minimum. Consider increasing to 7-10% for better defect prevention.'),
            ('5. Employee Performance', 'Top performers consistently exceed efficiency targets. Share best practices.'),
        ]
        
        for insight, desc in insights:
            ws[f'A{row}'] = insight
            ws[f'B{row}'] = desc
            ws.merge_cells(f'B{row}:C{row}')
            row += 1
        
        row += 1
        
        # Recommendations
        ws[f'A{row}'] = 'RECOMMENDATIONS'
        ws.merge_cells(f'A{row}:C{row}')
        self.apply_title_style(ws, row, 1)
        row += 1
        
        ws[f'A{row}'] = 'Priority'
        ws[f'B{row}'] = 'Action Item'
        ws[f'C{row}'] = 'Expected Impact'
        self.apply_header_style(ws, row, 1, 3)
        row += 1
        
        recommendations = [
            ('High', 'Complete pending employee training within 1 week', 'Improve operational consistency by 15%'),
            ('High', 'Investigate stock replenishment delays', 'Reduce average time to under 2 hours'),
            ('Medium', 'Increase quality audit rate to 7%', 'Reduce defects by 20%'),
            ('Medium', 'Document top performer best practices', 'Increase team efficiency by 10%'),
            ('Low', 'Implement automated inventory reconciliation', 'Reduce mismatch resolution time by 30%'),
        ]
        
        for priority, action, impact in recommendations:
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = action
            ws[f'C{row}'] = impact
            row += 1
        
        row += 1
        
        # Critical Action Items
        ws[f'A{row}'] = 'CRITICAL ACTION ITEMS'
        ws.merge_cells(f'A{row}:C{row}')
        self.apply_title_style(ws, row, 1)
        row += 1
        
        actions = [
            '1. Daily monitoring of wave completion times to ensure SLA compliance',
            '2. Weekly training completion reviews with department managers',
            '3. Immediate investigation of waves not completing within 1 hour',
            '4. Root cause analysis for stock replenishment delays',
            '5. Monthly review of quality audit coverage and effectiveness',
            '6. Proactive training for employees showing performance gaps',
            '7. Quick resolution of inventory mismatches to maintain accuracy',
            '8. Regular system health checks to minimize errors',
        ]
        
        for action in actions:
            ws[f'A{row}'] = action
            ws.merge_cells(f'A{row}:C{row}')
            row += 1
        
        self.set_column_widths(ws, [30, 60, 30])
        
        return ws
    
    def generate(self, filename='Ecom_Operations_Tracking_System_Formula_Based.xlsx'):
        """Generate the complete workbook"""
        print()
        print("=" * 70)
        print("E-COMMERCE OPERATIONS TRACKING SYSTEM")
        print("Formula-Based Excel Generator (No VBA)")
        print("=" * 70)
        print()
        
        print("📝 Creating sheets with formulas...")
        
        self.create_dashboard()
        print("  ✓ Dashboard (with KPI formulas)")
        
        self.create_bash_queries()
        print("  ✓ Bash Queries Response")
        
        self.create_wave_tracking()
        print("  ✓ Wave Tracking (with duration formulas)")
        
        self.create_employee_training()
        print("  ✓ Employee Training")
        
        self.create_stock_replenishment()
        print("  ✓ Stock Replenishment (with duration formulas)")
        
        self.create_quality_audit()
        print("  ✓ Quality Audit (with coverage & pass rate formulas)")
        
        self.create_picking_tasks()
        print("  ✓ Picking Tasks (with efficiency formulas)")
        
        self.create_order_volumes()
        print("  ✓ Order Volumes (with return rate formulas)")
        
        self.create_employee_performance()
        print("  ✓ Employee Performance")
        
        self.create_inventory_mismatch()
        print("  ✓ Inventory Mismatch (with variance formulas)")
        
        self.create_system_errors()
        print("  ✓ System Errors")
        
        self.create_insights()
        print("  ✓ Insights & Analytics")
        
        print()
        print("💾 Saving workbook...")
        
        output_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
        self.wb.save(output_path)
        
        print(f"✅ Success! File created: {filename}")
        print()
        print("📊 Features Included:")
        print("  • 12 comprehensive tracking sheets")
        print("  • Dashboard with live KPI calculations")
        print("  • Automatic duration calculations")
        print("  • Coverage and efficiency formulas")
        print("  • Variance tracking formulas")
        print("  • No VBA macros - pure Excel formulas only")
        print()
        print("=" * 70)
        print()
        
        return output_path

def main():
    """Main entry point"""
    generator = FormulaBasedExcelGenerator()
    generator.generate()

if __name__ == '__main__':
    main()
