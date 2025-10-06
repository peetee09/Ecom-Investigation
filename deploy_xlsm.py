#!/usr/bin/env python3
"""
Deployment Script for E-Commerce Operations Tracking System

This script creates a fully deployment-ready .xlsm file with VBA code embedded.
It automates the complete deployment process.

Usage:
    python3 deploy_xlsm.py                    # Create .xlsm without VBA (needs manual import)
    python3 deploy_xlsm.py --with-vba         # Create .xlsm with VBA (requires vbaProject.bin)
    python3 deploy_xlsm.py --extract <file>   # Extract VBA from existing .xlsm
"""

import xlsxwriter
import openpyxl
import os
import sys
import argparse
import zipfile
from datetime import datetime

class XLSMDeployer:
    """Creates deployment-ready .xlsm files with optional VBA embedding"""
    
    def __init__(self, base_dir=None):
        self.base_dir = base_dir or os.path.dirname(os.path.abspath(__file__))
        self.source_xlsx = os.path.join(self.base_dir, 'Ecom_Operations_Tracking_System.xlsx')
        self.output_xlsm = os.path.join(self.base_dir, 'Ecom_Operations_Tracking_System.xlsm')
        self.vba_bin = os.path.join(self.base_dir, 'vbaProject.bin')
        self.vba_module = os.path.join(self.base_dir, 'VBA_Modules.bas')
        self.vba_form = os.path.join(self.base_dir, 'BashQueryForm.frm')
    
    def print_header(self):
        """Print script header"""
        print()
        print("=" * 70)
        print("E-COMMERCE OPERATIONS TRACKING SYSTEM")
        print("Deployment-Ready XLSM Generator")
        print("=" * 70)
        print()
    
    def check_files(self):
        """Check if required files exist"""
        print("📋 Checking required files...")
        
        if not os.path.exists(self.source_xlsx):
            print(f"❌ Error: Source Excel file not found:")
            print(f"   {self.source_xlsx}")
            return False
        print(f"✓ Source Excel file: {os.path.basename(self.source_xlsx)}")
        
        if not os.path.exists(self.vba_module):
            print(f"⚠ Warning: VBA module not found:")
            print(f"   {self.vba_module}")
        else:
            print(f"✓ VBA module file: {os.path.basename(self.vba_module)}")
        
        if not os.path.exists(self.vba_form):
            print(f"ℹ Info: VBA form not found (optional):")
            print(f"   {self.vba_form}")
        else:
            print(f"✓ VBA form file: {os.path.basename(self.vba_form)}")
        
        print()
        return True
    
    def create_xlsm(self, include_vba=False):
        """
        Create the .xlsm file with optional VBA embedding
        
        Args:
            include_vba: If True, embed VBA from vbaProject.bin
        
        Returns:
            bool: True if successful
        """
        print("📝 Creating macro-enabled workbook...")
        print()
        
        # Check for VBA binary if requested
        has_vba = False
        if include_vba:
            if os.path.exists(self.vba_bin):
                has_vba = True
                print(f"✓ VBA binary found: {os.path.basename(self.vba_bin)}")
            else:
                print(f"⚠ Warning: VBA binary not found: {os.path.basename(self.vba_bin)}")
                print(f"  Creating .xlsm without embedded VBA.")
                print(f"  See instructions below for adding VBA manually.")
        
        # Read source workbook
        try:
            wb_source = openpyxl.load_workbook(self.source_xlsx, data_only=False)
            print(f"✓ Loaded source workbook: {len(wb_source.sheetnames)} sheets")
        except Exception as e:
            print(f"❌ Error loading source workbook: {e}")
            return False
        
        # Create new workbook
        try:
            workbook = xlsxwriter.Workbook(self.output_xlsm, {'constant_memory': False})
            
            # Add VBA project if available
            if has_vba:
                print(f"✓ Embedding VBA project...")
                workbook.add_vba_project(self.vba_bin)
                print(f"  VBA code is now embedded in the workbook!")
            
            # Create cell formats
            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#4472C4',
                'font_color': 'white',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })
            
            title_format = workbook.add_format({
                'bold': True,
                'font_size': 16,
                'bg_color': '#D9E1F2',
                'align': 'left'
            })
            
            subtitle_format = workbook.add_format({
                'italic': True,
                'font_size': 10,
                'font_color': '#666666'
            })
            
            # Copy sheets
            print()
            print("📑 Copying sheets:")
            for sheet_name in wb_source.sheetnames:
                print(f"  → {sheet_name}")
                ws_source = wb_source[sheet_name]
                ws_dest = workbook.add_worksheet(sheet_name)
                
                # Copy cell values
                max_row = 0
                max_col = 0
                for row_idx, row in enumerate(ws_source.iter_rows()):
                    for col_idx, cell in enumerate(row):
                        if cell.value is not None:
                            max_row = max(max_row, row_idx)
                            max_col = max(max_col, col_idx)
                            
                            # Apply formatting based on position
                            if row_idx == 0 and sheet_name == 'Dashboard':
                                ws_dest.write(row_idx, col_idx, cell.value, title_format)
                            elif row_idx == 1 and sheet_name == 'Dashboard':
                                ws_dest.write(row_idx, col_idx, cell.value, subtitle_format)
                            elif row_idx == 0:
                                ws_dest.write(row_idx, col_idx, cell.value, header_format)
                            else:
                                ws_dest.write(row_idx, col_idx, cell.value)
                
                # Set column widths
                if sheet_name == 'Dashboard':
                    ws_dest.set_column('A:A', 45)
                    ws_dest.set_column('B:B', 15)
                    ws_dest.set_column('C:C', 15)
                    ws_dest.set_column('D:D', 20)
                    ws_dest.set_column('E:E', 20)
                elif 'Response' in sheet_name or 'Query' in sheet_name:
                    ws_dest.set_column('A:B', 12)
                    ws_dest.set_column('C:D', 20)
                    ws_dest.set_column('E:G', 25)
                    ws_dest.set_column('H:H', 30)
                else:
                    ws_dest.set_column('A:Z', 14)
                
                # Freeze first row for non-Dashboard sheets
                if sheet_name != 'Dashboard':
                    ws_dest.freeze_panes(1, 0)
            
            # Close workbook
            workbook.close()
            wb_source.close()
            
            print()
            print("=" * 70)
            print("✅ SUCCESS!")
            print("=" * 70)
            print()
            print(f"Created: {os.path.basename(self.output_xlsm)}")
            
            if os.path.exists(self.output_xlsm):
                size_kb = os.path.getsize(self.output_xlsm) / 1024
                print(f"Size: {size_kb:.1f} KB")
                print(f"Location: {self.output_xlsm}")
            
            if has_vba:
                print()
                print("🎉 VBA code is EMBEDDED and ready to use!")
                print()
                print("To use the file:")
                print("  1. Open Ecom_Operations_Tracking_System.xlsm")
                print("  2. Enable macros when prompted")
                print("  3. Press Alt + F8 to see available macros")
                print("  4. Run macros like ShowDashboard, AddBashQuery, etc.")
                print()
                print("The file is now DEPLOYMENT-READY! 🚀")
            else:
                print()
                print("⚠️  VBA code is NOT embedded yet.")
                self.print_vba_instructions()
            
            return True
            
        except Exception as e:
            print(f"❌ Error creating workbook: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def extract_vba(self, xlsm_file):
        """
        Extract vbaProject.bin from an existing .xlsm file
        
        Args:
            xlsm_file: Path to source .xlsm file
        
        Returns:
            bool: True if successful
        """
        print(f"📂 Extracting VBA from: {os.path.basename(xlsm_file)}")
        print()
        
        if not os.path.exists(xlsm_file):
            print(f"❌ Error: File not found: {xlsm_file}")
            return False
        
        try:
            with zipfile.ZipFile(xlsm_file, 'r') as zip_ref:
                vba_path = 'xl/vbaProject.bin'
                if vba_path in zip_ref.namelist():
                    vba_data = zip_ref.read(vba_path)
                    
                    with open(self.vba_bin, 'wb') as f:
                        f.write(vba_data)
                    
                    print(f"✅ Successfully extracted vbaProject.bin")
                    print(f"   Size: {len(vba_data):,} bytes")
                    print(f"   Saved to: {self.vba_bin}")
                    print()
                    print("✓ You can now run this script with --with-vba to create")
                    print("  a new .xlsm file with the extracted VBA code embedded.")
                    return True
                else:
                    print(f"❌ No VBA project found in {xlsm_file}")
                    return False
        except Exception as e:
            print(f"❌ Error extracting VBA: {e}")
            return False
    
    def print_vba_instructions(self):
        """Print instructions for manually adding VBA"""
        print()
        print("=" * 70)
        print("HOW TO ADD VBA CODE")
        print("=" * 70)
        print()
        print("To complete the setup, add VBA code manually:")
        print()
        print("1. Open Ecom_Operations_Tracking_System.xlsm in Excel")
        print("2. Press Alt + F11 to open VBA Editor")
        print("3. Go to File → Import File")
        print("4. Select VBA_Modules.bas and click Open")
        print("5. (Optional) Import BashQueryForm.frm for advanced features")
        print("6. Close VBA Editor (Alt + Q)")
        print("7. Save the file (Ctrl + S)")
        print()
        print("After adding VBA, you can re-extract it:")
        print(f"  python3 {os.path.basename(__file__)} --extract Ecom_Operations_Tracking_System.xlsm")
        print()
        print("Then recreate the file with embedded VBA:")
        print(f"  python3 {os.path.basename(__file__)} --with-vba")
        print()

def main():
    """Main entry point"""
    parser = argparse.ArgumentParser(
        description='Create deployment-ready .xlsm file with VBA code',
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument('--with-vba', action='store_true',
                        help='Embed VBA from vbaProject.bin')
    parser.add_argument('--extract', metavar='FILE',
                        help='Extract VBA from existing .xlsm file')
    
    args = parser.parse_args()
    
    deployer = XLSMDeployer()
    deployer.print_header()
    
    # Extract mode
    if args.extract:
        if deployer.extract_vba(args.extract):
            return 0
        else:
            return 1
    
    # Create mode
    if not deployer.check_files():
        return 1
    
    if deployer.create_xlsm(include_vba=args.with_vba):
        print()
        return 0
    else:
        print()
        print("❌ Failed to create .xlsm file")
        return 1

if __name__ == '__main__':
    sys.exit(main())
