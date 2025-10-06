#!/usr/bin/env python3
"""
E-commerce Operations Tracking System Generator
Creates a comprehensive Excel workbook for monitoring warehouse operations
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

def create_ecom_tracking_system():
    """Create comprehensive Excel tracking system for e-commerce operations"""
    
    wb = Workbook()
    
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Define common styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="D6DCE4", end_color="D6DCE4", fill_type="solid")
    subheader_font = Font(bold=True, size=10)
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 1. Dashboard / Summary Sheet
    create_dashboard_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border)
    
    # 2. Bash Queries Response Sheet
    create_bash_queries_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border)
    
    # 3. Wave Tracking Sheet
    create_wave_tracking_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 4. Employee Training Sheet
    create_employee_training_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 5. Stock Replenishment Sheet
    create_stock_replenishment_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 6. Quality Audit Sheet
    create_quality_audit_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 7. Picking Task Monitoring
    create_picking_task_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 8. Order Volumes Sheet
    create_order_volumes_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 9. Employee Performance Sheet
    create_employee_performance_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 10. Inventory Mismatch Sheet
    create_inventory_mismatch_sheet(wb, header_fill, header_font, center_alignment, thin_border)
    
    # 11. System Errors Sheet
    create_system_errors_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border)
    
    # 12. Insights & Analytics Sheet
    create_insights_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border)
    
    # Save the workbook
    filename = "Ecom_Operations_Tracking_System.xlsx"
    wb.save(filename)
    print(f"✓ Excel tracking system created successfully: {filename}")
    return filename


def create_dashboard_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border):
    """Create main dashboard with KPIs and status overview"""
    ws = wb.create_sheet("Dashboard", 0)
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "E-COMMERCE OPERATIONS DASHBOARD"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = center_alignment
    
    # Date
    ws.merge_cells('A2:H2')
    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    ws['A2'].alignment = center_alignment
    ws['A2'].font = Font(italic=True, size=10)
    
    # KPI Section
    ws['A4'] = "KEY PERFORMANCE INDICATORS"
    ws['A4'].font = Font(bold=True, size=12)
    ws.merge_cells('A4:H4')
    
    # KPI Headers
    headers = ['Metric', 'Target', 'Current', 'Status', 'Last Updated']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample KPIs
    kpis = [
        ['Wave Completion (1 hour)', '100%', '=COUNTA(\'Wave Tracking\'!A:A)-1', ''],
        ['Employee Training Completion', '100%', '=COUNTA(\'Employee Training\'!A:A)-1', ''],
        ['Picking Within Hours Target', '95%', '=COUNTA(\'Picking Tasks\'!A:A)-1', ''],
        ['Quality Audit Coverage', '>5%', '=COUNTA(\'Quality Audit\'!A:A)-1', ''],
        ['Open System Errors', '0', '=COUNTA(\'System Errors\'!A:A)-1', ''],
        ['Inventory Mismatches', '0', '=COUNTA(\'Inventory Mismatch\'!A:A)-1', ''],
        ['SLA Compliance Rate', '>98%', '98.5%', 'ON TARGET'],
    ]
    
    for row, kpi_data in enumerate(kpis, 6):
        for col, value in enumerate(kpi_data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            if col == 1:
                cell.alignment = left_alignment
            else:
                cell.alignment = center_alignment
        # Last Updated
        ws.cell(row=row, column=5).value = datetime.now().strftime('%Y-%m-%d %H:%M')
    
    # Quick Links Section
    ws['A15'] = "MONITORING AREAS"
    ws['A15'].font = Font(bold=True, size=12)
    ws.merge_cells('A15:H15')
    
    monitoring_areas = [
        '• Bash Queries Response - Track customer order queries',
        '• Wave Tracking - Monitor wave completion status',
        '• Employee Training - Track training completion',
        '• Stock Replenishment - Monitor stock refill times',
        '• Quality Audit - Track quality checks and rechecks',
        '• Picking Tasks - Monitor picking efficiency',
        '• Order Volumes - Track daily order volumes',
        '• Employee Performance - Monitor staff performance',
        '• Inventory Mismatch - Track inventory discrepancies',
        '• System Errors - Log and track system issues',
    ]
    
    for idx, area in enumerate(monitoring_areas, 16):
        ws[f'A{idx}'] = area
        ws[f'A{idx}'].font = Font(size=10)
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20


def create_bash_queries_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border):
    """Create sheet for responding to Bash queries about customer orders"""
    ws = wb.create_sheet("Bash Queries Response")
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "BASH QUERIES - CUSTOMER ORDER RESPONSES"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Query ID', 'Date/Time', 'Customer Name', 'Order ID', 'Query Type', 
               'Query Details', 'Status', 'Response', 'Responded By', 'Response Time (mins)']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    sample_queries = [
        ['Q001', datetime.now().strftime('%Y-%m-%d %H:%M'), 'John Doe', 'ORD-10001', 'Order Status',
         'Where is my order?', 'RESPONDED', 'Order is in picking stage', 'Agent A', '5'],
        ['Q002', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Jane Smith', 'ORD-10002', 'Delivery Time',
         'When will order arrive?', 'RESPONDED', 'Expected delivery tomorrow', 'Agent B', '3'],
        ['Q003', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Mike Johnson', 'ORD-10003', 'Order Issue',
         'Wrong item received', 'IN PROGRESS', 'Investigating with warehouse', 'Agent A', ''],
    ]
    
    for row, query in enumerate(sample_queries, 3):
        for col, value in enumerate(query, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment if col != 6 else left_alignment
    
    # Set column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['H'].width = 30


def create_wave_tracking_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for tracking wave completion status"""
    ws = wb.create_sheet("Wave Tracking")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "WAVE TRACKING - COMPLETION STATUS"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Wave ID', 'Date', 'Start Time', 'Expected Completion', 
               'Actual Completion', 'Status', 'Orders Count', 'Completion Time (mins)', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    now = datetime.now()
    sample_waves = [
        ['WAVE-001', now.strftime('%Y-%m-%d'), '08:00', '09:00', '08:55', 'COMPLETED', '150', '55', 'On time'],
        ['WAVE-002', now.strftime('%Y-%m-%d'), '09:00', '10:00', '10:05', 'COMPLETED', '145', '65', 'Slight delay'],
        ['WAVE-003', now.strftime('%Y-%m-%d'), '10:00', '11:00', '', 'IN PROGRESS', '160', '', ''],
    ]
    
    for row, wave in enumerate(sample_waves, 3):
        for col, value in enumerate(wave, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['I'].width = 25


def create_employee_training_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for tracking employee training"""
    ws = wb.create_sheet("Employee Training")
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "EMPLOYEE TRAINING RECORDS"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Employee ID', 'Employee Name', 'Department', 'Training Type', 
               'Scheduled Date', 'Completion Date', 'Status', 'Score/Result', 'Trainer', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    sample_training = [
        ['EMP001', 'John Smith', 'Warehouse', 'Picking Training', '2024-01-15', '2024-01-15', 'COMPLETED', '95%', 'Trainer A', 'Excellent'],
        ['EMP002', 'Sarah Wilson', 'Warehouse', 'Quality Audit', '2024-01-16', '2024-01-16', 'COMPLETED', '88%', 'Trainer B', 'Good'],
        ['EMP003', 'Mike Brown', 'Warehouse', 'Safety Training', '2024-01-20', '', 'SCHEDULED', '', 'Trainer A', ''],
        ['EMP004', 'Lisa Davis', 'Packing', 'Packing Standards', '2024-01-18', '2024-01-18', 'COMPLETED', '92%', 'Trainer C', 'Very good'],
    ]
    
    for row, training in enumerate(sample_training, 3):
        for col, value in enumerate(training, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['J'].width = 25


def create_stock_replenishment_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for monitoring stock replenishment"""
    ws = wb.create_sheet("Stock Replenishment")
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "STOCK REPLENISHMENT MONITORING"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Replenishment ID', 'Item SKU', 'Item Name', 'Current Stock', 'Reorder Point',
               'Quantity Ordered', 'Request Time', 'Completion Time', 'Duration (hours)', 
               'Status', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    now = datetime.now()
    sample_replenishment = [
        ['REP-001', 'SKU-1001', 'Widget A', '50', '100', '500', 
         (now - timedelta(hours=3)).strftime('%Y-%m-%d %H:%M'), 
         (now - timedelta(hours=1)).strftime('%Y-%m-%d %H:%M'), '2.0', 'COMPLETED', 'On time'],
        ['REP-002', 'SKU-1002', 'Widget B', '25', '100', '750', 
         (now - timedelta(hours=5)).strftime('%Y-%m-%d %H:%M'), 
         now.strftime('%Y-%m-%d %H:%M'), '5.0', 'COMPLETED', 'Delayed - investigate'],
        ['REP-003', 'SKU-1003', 'Widget C', '10', '100', '1000', 
         (now - timedelta(hours=2)).strftime('%Y-%m-%d %H:%M'), 
         '', '', 'IN PROGRESS', 'Urgent'],
    ]
    
    for row, replen in enumerate(sample_replenishment, 3):
        for col, value in enumerate(replen, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['K'].width = 25


def create_quality_audit_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for quality audit tracking"""
    ws = wb.create_sheet("Quality Audit")
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "QUALITY AUDIT TRACKING (Target: >5% Diversion)"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Audit ID', 'Date', 'Order ID', 'Item SKU', 'Audit Type', 
               'Result', 'Issue Found', 'Action Taken', 'Audited By', 'Recheck Required', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Add statistics row
    ws['A3'] = "STATISTICS:"
    ws['B3'] = "Total Orders Today:"
    ws['C3'] = 1000
    ws['D3'] = "Audited (Current %):"
    ws['E3'] = '=COUNTA(A5:A100)'
    ws['F3'] = "Audit %:"
    ws['G3'] = '=(E3/C3)*100'
    ws['H3'] = "Target: >5%"
    
    # Sample data
    sample_audits = [
        ['AUD-001', datetime.now().strftime('%Y-%m-%d'), 'ORD-10001', 'SKU-1001', 'Random Check',
         'PASS', 'None', 'None', 'QA Team A', 'NO', ''],
        ['AUD-002', datetime.now().strftime('%Y-%m-%d'), 'ORD-10015', 'SKU-1002', 'Random Check',
         'FAIL', 'Wrong quantity', 'Repack ordered', 'QA Team B', 'YES', 'Critical'],
        ['AUD-003', datetime.now().strftime('%Y-%m-%d'), 'ORD-10025', 'SKU-1003', 'Random Check',
         'PASS', 'None', 'None', 'QA Team A', 'NO', ''],
        ['AUD-004', datetime.now().strftime('%Y-%m-%d'), 'ORD-10042', 'SKU-1004', 'Targeted Check',
         'FAIL', 'Damaged item', 'Item replaced', 'QA Team C', 'YES', 'Minor damage'],
    ]
    
    for row, audit in enumerate(sample_audits, 5):
        for col, value in enumerate(audit, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['G'].width = 22
    ws.column_dimensions['K'].width = 25


def create_picking_task_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for picking task monitoring"""
    ws = wb.create_sheet("Picking Tasks")
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "PICKING TASK MONITORING"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Task ID', 'Wave ID', 'Order ID', 'Picker ID', 'Start Time', 
               'End Time', 'Duration (mins)', 'Items Count', 'Status', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    now = datetime.now()
    sample_picking = [
        ['PICK-001', 'WAVE-001', 'ORD-10001', 'EMP001', '08:00', '08:15', '15', '5', 'COMPLETED', ''],
        ['PICK-002', 'WAVE-001', 'ORD-10002', 'EMP002', '08:05', '08:25', '20', '8', 'COMPLETED', ''],
        ['PICK-003', 'WAVE-002', 'ORD-10003', 'EMP001', '09:00', '09:30', '30', '12', 'COMPLETED', ''],
        ['PICK-004', 'WAVE-003', 'ORD-10004', 'EMP003', '10:00', '', '', '6', 'IN PROGRESS', ''],
    ]
    
    for row, pick in enumerate(sample_picking, 3):
        for col, value in enumerate(pick, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['J'].width = 25


def create_order_volumes_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for order volume tracking"""
    ws = wb.create_sheet("Order Volumes")
    
    # Title
    ws.merge_cells('A1:I1')
    ws['A1'] = "ORDER VOLUMES MONITORING"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Date', 'Total Orders', 'Orders Picked', 'Orders Packed', 'Orders Shipped',
               'Pending Orders', 'Cancelled Orders', 'Average Processing Time (hrs)', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data for last 7 days
    for day in range(7, 0, -1):
        date = datetime.now() - timedelta(days=day)
        row = 10 - day
        data = [
            date.strftime('%Y-%m-%d'),
            950 + (day * 10),
            940 + (day * 10),
            935 + (day * 10),
            930 + (day * 10),
            20 - day,
            5,
            2.5,
            'Normal operations'
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 20
    ws.column_dimensions['I'].width = 25


def create_employee_performance_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for employee performance tracking"""
    ws = wb.create_sheet("Employee Performance")
    
    # Title
    ws.merge_cells('A1:K1')
    ws['A1'] = "EMPLOYEE PERFORMANCE METRICS"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Employee ID', 'Employee Name', 'Department', 'Date', 'Tasks Completed',
               'Average Time (mins)', 'Accuracy %', 'Training Status', 'Performance Rating', 
               'Improvement Areas', 'Notes']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    sample_performance = [
        ['EMP001', 'John Smith', 'Picking', datetime.now().strftime('%Y-%m-%d'), 
         '45', '18', '98%', 'COMPLETED', 'EXCELLENT', 'None', 'Top performer'],
        ['EMP002', 'Sarah Wilson', 'Picking', datetime.now().strftime('%Y-%m-%d'), 
         '42', '20', '96%', 'COMPLETED', 'GOOD', 'Speed', 'Consistent'],
        ['EMP003', 'Mike Brown', 'Packing', datetime.now().strftime('%Y-%m-%d'), 
         '38', '22', '94%', 'IN PROGRESS', 'AVERAGE', 'Accuracy', 'Needs training'],
        ['EMP004', 'Lisa Davis', 'Packing', datetime.now().strftime('%Y-%m-%d'), 
         '48', '17', '99%', 'COMPLETED', 'EXCELLENT', 'None', 'Star performer'],
    ]
    
    for row, perf in enumerate(sample_performance, 3):
        for col, value in enumerate(perf, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 12):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['J'].width = 22
    ws.column_dimensions['K'].width = 25


def create_inventory_mismatch_sheet(wb, header_fill, header_font, center_alignment, thin_border):
    """Create sheet for inventory mismatch tracking"""
    ws = wb.create_sheet("Inventory Mismatch")
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "INVENTORY MISMATCH TRACKING"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Mismatch ID', 'Date Discovered', 'Item SKU', 'Item Name', 'System Count',
               'Physical Count', 'Variance', 'Root Cause', 'Resolution', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    sample_mismatches = [
        ['MIS-001', datetime.now().strftime('%Y-%m-%d'), 'SKU-1001', 'Widget A', 
         '500', '498', '-2', 'Picking error', 'System updated', 'RESOLVED'],
        ['MIS-002', datetime.now().strftime('%Y-%m-%d'), 'SKU-1005', 'Widget E', 
         '200', '205', '+5', 'Receiving error', 'Under investigation', 'IN PROGRESS'],
        ['MIS-003', datetime.now().strftime('%Y-%m-%d'), 'SKU-1008', 'Widget H', 
         '150', '148', '-2', 'Unknown', 'Investigating', 'OPEN'],
    ]
    
    for row, mismatch in enumerate(sample_mismatches, 3):
        for col, value in enumerate(mismatch, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # Set column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['D'].width = 22
    ws.column_dimensions['H'].width = 22


def create_system_errors_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border):
    """Create sheet for system errors logging"""
    ws = wb.create_sheet("System Errors")
    
    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "SYSTEM ERRORS LOG"
    ws['A1'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A1'].fill = header_fill
    ws['A1'].alignment = center_alignment
    
    # Headers
    headers = ['Error ID', 'Date/Time', 'System/Module', 'Error Type', 'Severity',
               'Description', 'Impact', 'Resolution', 'Resolved By', 'Status']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = thin_border
    
    # Sample data
    sample_errors = [
        ['ERR-001', datetime.now().strftime('%Y-%m-%d %H:%M'), 'WMS', 'Database Connection', 'HIGH',
         'Connection timeout', 'Picking delayed 5 mins', 'Server restarted', 'IT Team', 'RESOLVED'],
        ['ERR-002', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Barcode Scanner', 'Hardware', 'MEDIUM',
         'Scanner not reading', 'Manual entry required', 'Scanner replaced', 'IT Team', 'RESOLVED'],
        ['ERR-003', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Packing System', 'Software Bug', 'LOW',
         'Print label delay', 'Minor slowdown', 'Investigating', 'IT Team', 'OPEN'],
    ]
    
    for row, error in enumerate(sample_errors, 3):
        for col, value in enumerate(error, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = value
            cell.border = thin_border
            cell.alignment = center_alignment if col not in [6, 7, 8] else left_alignment
    
    # Set column widths
    for col in range(1, 11):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 25


def create_insights_sheet(wb, header_fill, header_font, center_alignment, left_alignment, thin_border):
    """Create sheet for insights and analytics"""
    ws = wb.create_sheet("Insights & Analytics")
    
    # Title
    ws.merge_cells('A1:H1')
    ws['A1'] = "INSIGHTS & ANALYTICS"
    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A1'].alignment = center_alignment
    
    # Outcomes Section
    ws['A3'] = "EXPECTED OUTCOMES"
    ws['A3'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A3'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.merge_cells('A3:H3')
    
    outcomes = [
        ['✓ People\'s Performance Improvement', 'Track individual and team metrics to identify training needs and improvement areas'],
        ['✓ SLA Met Continuously', 'Monitor service level agreements and ensure consistent compliance'],
        ['✓ Operations Optimised', 'Analyze bottlenecks and streamline processes for efficiency'],
    ]
    
    for idx, outcome in enumerate(outcomes, 4):
        ws[f'A{idx}'] = outcome[0]
        ws[f'B{idx}'] = outcome[1]
        ws.merge_cells(f'B{idx}:H{idx}')
        ws[f'A{idx}'].font = Font(bold=True)
    
    # Key Insights Section
    ws['A8'] = "KEY INSIGHTS"
    ws['A8'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A8'].fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    ws.merge_cells('A8:H8')
    
    insights = [
        ['Wave Completion', 'Monitor waves completing within 1 hour target. Identify patterns of delays.'],
        ['Training Effectiveness', 'Track training completion and correlate with performance improvements.'],
        ['Replenishment Speed', 'Analyze stock replenishment times to reduce picking delays.'],
        ['Quality Coverage', 'Ensure >5% quality audit coverage and track defect rates.'],
        ['Employee Productivity', 'Monitor picking/packing rates and identify top performers for recognition.'],
        ['Error Patterns', 'Track system errors and inventory mismatches to prevent recurrence.'],
        ['Customer Satisfaction', 'Fast response to Bash queries improves customer satisfaction.'],
    ]
    
    for idx, insight in enumerate(insights, 9):
        ws[f'A{idx}'] = insight[0]
        ws[f'B{idx}'] = insight[1]
        ws.merge_cells(f'B{idx}:H{idx}')
        ws[f'A{idx}'].font = Font(bold=True)
        ws[f'A{idx}'].alignment = left_alignment
        ws[f'B{idx}'].alignment = left_alignment
    
    # Action Items Section
    ws['A17'] = "RECOMMENDED ACTIONS"
    ws['A17'].font = Font(bold=True, size=12, color="FFFFFF")
    ws['A17'].fill = PatternFill(start_color="5B9BD5", end_color="5B9BD5", fill_type="solid")
    ws.merge_cells('A17:H17')
    
    actions = [
        '1. Daily review of dashboard KPIs to catch issues early',
        '2. Weekly analysis of employee performance trends',
        '3. Immediate investigation of waves not completing within 1 hour',
        '4. Root cause analysis for stock replenishment delays',
        '5. Monthly review of quality audit coverage and effectiveness',
        '6. Proactive training for employees showing performance gaps',
        '7. Quick resolution of inventory mismatches to maintain accuracy',
        '8. Regular system health checks to minimize errors',
    ]
    
    for idx, action in enumerate(actions, 18):
        ws[f'A{idx}'] = action
        ws.merge_cells(f'A{idx}:H{idx}')
        ws[f'A{idx}'].alignment = left_alignment
    
    # Set column widths
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 60


if __name__ == "__main__":
    print("=" * 60)
    print("E-Commerce Operations Tracking System Generator")
    print("=" * 60)
    create_ecom_tracking_system()
    print("\nSystem ready for use!")
    print("=" * 60)
